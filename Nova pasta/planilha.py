import openpyxl

#criar planilha
book = openpyxl.Workbook()

#como visualizar paginas existentes
print(book.sheetnames)

#como criar uma pagina 
book.create_sheet("Computadores")
#como selecionar uma pagina
frutas_page = book['Computadores']
frutas_page.append(["Eletrônica", "Memória ram", "preço"])
frutas_page.append(['Computador 1', '8gb ram', 'R$ 2500'])
frutas_page.append(['Computador 2', '16gb ram', 'R$ 5500'])
frutas_page.append(['Computador 3', '32gb ram','R$ 8500'])
# salvar planilha
book.save('Meuscomputadores.xlsx')